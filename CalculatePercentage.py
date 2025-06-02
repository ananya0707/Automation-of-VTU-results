import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class PassPercentageCalculator:
    def __init__(self, file_path):
        """
        Initialize the class with the Excel file path.
        """
        self.file_path = file_path
        self.wb = None
        self.sheet = None
        self.df = None

    def load_file(self):
        """
        Load the Excel file and initialize workbook and sheet.
        """
        self.wb = load_workbook(self.file_path)
        self.sheet = self.wb.active
        self.df = pd.read_excel(self.file_path)

    def calculate_pass_percentage(self):
        """
        Calculate the pass percentage for each 'Result' column and append it to the file.
        """
        # Ensure the file is loaded
        if self.df is None or self.sheet is None:
            raise ValueError("File not loaded. Call 'load_file' first.")

        # Identify columns ending with "Result"
        result_columns = [col for col in self.df.columns if col.endswith("Result")]

        # Determine the row where percentages will be added (leave one blank row)
        new_row_index = len(self.df) + 3

        # Write "Percentage Passed" in the first column of the percentage row
        self.sheet.cell(row=new_row_index, column=1).value = "Percentage Passed"
        self.sheet.cell(row=new_row_index, column=1).alignment = Alignment(horizontal="center", vertical="center")

        # Append pass percentages for each "Result" column
        for col in result_columns:
            total_students = len(self.df)
            passed_students = self.df[self.df[col] == "P"].shape[0]
            pass_percentage = round((passed_students / total_students) * 100, 2)  # Format as float with 2 decimals

            # Find the column index for the "Result" column
            col_index = self.df.columns.get_loc(col) + 1  # Adjust for Excel's 1-based indexing
            percentage_cell = self.sheet.cell(row=new_row_index, column=col_index)
            percentage_cell.value = pass_percentage
            percentage_cell.alignment = Alignment(horizontal="center", vertical="center")

    def save_file(self):
        """
        Save the updated Excel file to the specified output path.
        """
        if self.wb is None:
            raise ValueError("Workbook not initialized. Call 'load_file' first.")
        self.wb.save(self.file_path)
        print(f"Updated file saved to: {self.file_path}")
